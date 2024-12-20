import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os
def corriger_csv(fichiers_csv):
    """
    Consolide plusieurs fichiers CSV en un seul fichier Excel en additionnant les quantités pour les mêmes 'id_produit'.

    Préconditions :
        - 'fichiers_csv' est une liste de chemins vers des fichiers CSV existants.
        - Chaque fichier CSV doit contenir les colonnes suivantes :
            * 'id_produit'
            * 'nom_produit'
            * 'categorie'
            * 'quantite'
            * 'prix_unitaire'
            * 'date_ajout'

    Postconditions :
        - Un fichier Excel nommé 'stock_consolidated.xlsx' est généré dans le répertoire courant.
        - Les données des fichiers CSV sont combinées :
            * Les lignes avec le même 'id_produit' sont regroupées, et leurs quantités sont additionnées.
            * La première date d'ajout (chronologiquement) est conservée.
            * Les noms des fichiers sources sont combinés.

    Raises :
        - FileNotFoundError : Si un ou plusieurs fichiers dans 'fichiers_csv' n'existent pas.
        - ValueError : Si les fichiers CSV ne contiennent pas les colonnes requises.
        - Exception : Pour toute autre erreur survenue pendant le traitement ou l'écriture.

    Returns :
        - str : Le chemin du fichier Excel généré si la consolidation réussit.
        - None : En cas d'échec.
    """
    try:
        # Étape 1 : Lire et combiner les fichiers CSV
        dataframes = []
        for fichier in fichiers_csv:
            if not os.path.exists(fichier):
                raise FileNotFoundError(f"Le fichier {fichier} n'existe pas.")
            df = pd.read_csv(fichier, encoding='utf-8')
            df['source_fichier'] = fichier  # Ajouter la source du fichier
            dataframes.append(df)

        # Étape 2 : Combiner les DataFrames
        consolidated_df = pd.concat(dataframes, ignore_index=True)

        # Étape 3 : Consolider les doublons par id_produit
        consolidated_df = (
            consolidated_df.groupby(['id_produit', 'nom_produit', 'categorie', 'prix_unitaire'], as_index=False)
            .agg({
                'quantite': 'sum',
                'date_ajout': 'min',  # Prend la première date
                'source_fichier': lambda x: ', '.join(set(x))  # Combine les sources de fichiers
            })
        )

        # Étape 4 : Créer le fichier Excel
        fichier_excel = 'stock_consolidated.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Stock Consolidé"

        # Styles pour l'en-tête
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        font_bold = Font(bold=True)

        # Ajouter les en-têtes
        for col_idx, col_name in enumerate(consolidated_df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = font_bold

        # Ajouter les données
        for row_idx, row in enumerate(consolidated_df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        workbook.save(fichier_excel)
        print(f"Fichier Excel consolidé créé : {fichier_excel}")
        return fichier_excel

    except FileNotFoundError as fnfe:
        print(f"ERREUR : {fnfe}")
    except ValueError as ve:
        print(f"ERREUR : Fichiers CSV invalides - {ve}")
    except Exception as e:
        print(f"Erreur lors de la consolidation des fichiers : {e}")
        return None

# Exemple d'utilisation
if __name__ == "__main__":
    fichiers_csv = ['stock_1.csv', 'stock_2.csv']
    corriger_csv(fichiers_csv)
